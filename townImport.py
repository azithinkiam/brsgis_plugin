import openpyxl

path = "D:\\LeightonProjects\\GEO DATA\\Town_Data\\BBH -\\BBHExport.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# cell_obj = sheet_obj.cell(row=1, column=1)

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
planData = []
allPlans = []
error_count = 0

for j in range(2, max_row + 1):

    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=j, column=i)
        # QgsMessageLog.logMessage(str(j) + ' ' + str(i), 'BRS_GIS', level=Qgis.Info)
        planData.append(str(cell_obj.value))
        # QgsMessageLog.logMessage(str(planData), 'BRS_GIS', level=Qgis.Info)
    allPlans.append(planData)
    planData = []

for plan in allPlans:

    # QgsMessageLog.logMessage(str(plan), 'BRS_GIS', level=Qgis.Info)

    map_bk_lot = str(plan[0])
    if map_bk_lot == '':
        pass
    else:
        # QgsMessageLog.logMessage(map_bk_lot, 'BRS_GIS', level=Qgis.Info)
        location = plan[1]
        name = plan[2]
        mailingAddress = plan[3]
        deeds = plan[5]
        pages = deeds.split(" ")
        numDeeds = len(pages)
        # QgsMessageLog.logMessage(map_bk_lot + ' - TOTAL DEED VALUES: ' + str(numDeeds), 'BRS_GIS', level=Qgis.Info)
        # QgsMessageLog.logMessage('PAGES: ' + str(pages), 'BRS_GIS', level=Qgis.Info)

        for x in range(0, numDeeds):
            if x == 0:
                bookpage = str(pages[x])
            elif x == 1:
                bookpage = bookpage + ',' + str(pages[x])

        addressName = plan[10]
        addressStreet = plan[11]
        addressStreet2 = plan[12]
        addressStreet3 = plan[13]
        addressStreet4 = plan[14]
        addressCity = plan[15]
        addressState = plan[16]
        addressZip = plan[17]

        QgsMessageLog.logMessage(map_bk_lot + ' | DEED TO KEEP: ' + bookpage + ' | ALL DEED INFO: ' + str(pages),
                                 'BRS_GIS', level=Qgis.Info)
